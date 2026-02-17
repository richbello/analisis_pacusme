import os
import re
import io
import tempfile
import traceback
from datetime import datetime, timedelta
from typing import Tuple, Dict, Optional
import pandas as pd
import numpy as np
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURACIÓN DE USUARIOS Y ROLES
# ============================================================
USUARIOS = {
    "admin": {
        "password": "admin123",
        "rol": "Administrador",
        "permisos": ["seguridad", "analisis_pac"]
    },
    "contador": {
        "password": "contador2025",
        "rol": "Contador",
        "permisos": ["seguridad", "analisis_pac"]
    },
    "auxiliar": {
        "password": "aux123",
        "rol": "Auxiliar Contable",
        "permisos": ["seguridad", "analisis_pac"]
    },
    "auditor": {
        "password": "audit2025",
        "rol": "Auditor",
        "permisos": ["seguridad", "analisis_pac"]
    },
    "jefe": {
        "password": "jefe2025",
        "rol": "Jefe de Área",
        "permisos": ["seguridad", "analisis_pac"]
    }
}

# ============================================================
# COLORES Y ESTILOS - BANDERA DE BOGOTÁ
# ============================================================
BOGOTA_YELLOW = "#FCDD09"
BOGOTA_RED = "#DA121A"
DARK = "#111827"
WHITE = "#FFFFFF"
BLACK = "#111111"

# ============================================================
# CONFIGURACIÓN DE SEGURIDAD
# ============================================================
MAX_INTENTOS = 5
BLOQUEO_MINUTOS = 3
MOSTRAR_INTENTOS = True

# ============================================================
# CLASE ANALIZADOR PAC
# ============================================================

class AnalizadorPAC:
    
    def __init__(self, df):
        self.df = df
        self.resumen_periodo = None
        self.resumen_centro = None
        self.resumen_fondos = None
        self.resumen_detallado = None
        self.analisis = {}
        
    def limpiar_datos(self):
        """Limpia datos eliminando filas totales y NaN"""
        self.df = self.df[self.df['Per.presup.'].notna()]
        self.df = self.df[self.df['Centro gestor'].notna()]
        return self
    
    def resumen_detallado_completo(self):
        """Genera resumen detallado mostrando cada registro individual"""
        
        # Definir las columnas que queremos mostrar
        columnas_mostrar = []
        columnas_renombrar = {}
        
        # Mapeo de columnas originales a nombres finales
        mapeo_columnas = {
            'Progr.financiación': 'Programa Financiación',
            'Centro gestor': 'Centro Gestor',
            'Per.presup.': 'Período',
            'Pos.presupuestaria': 'Pos.Presupuestaria',
            'Fondos': 'Fondos',
            'PAC inicial': 'PAC Inicial',
            'PAC Actual': 'PAC Actual',
            'PAC Reprogramado': 'PAC Reprogramado',
            'Girado y Recaudado PAC': 'Girado y Recaudado',
            'Disponibilidad PAC': 'Disponibilidad'
        }
        
        # Verificar qué columnas existen en el DataFrame
        for col_original, col_nueva in mapeo_columnas.items():
            if col_original in self.df.columns:
                columnas_mostrar.append(col_original)
                columnas_renombrar[col_original] = col_nueva
        
        if not columnas_mostrar:
            st.warning("⚠️ No se encontraron columnas para mostrar")
            return pd.DataFrame()
        
        # Seleccionar solo las columnas que existen
        self.resumen_detallado = self.df[columnas_mostrar].copy()
        
        # Calcular % Ejecutado
        if 'PAC Actual' in self.resumen_detallado.columns and 'Girado y Recaudado PAC' in self.resumen_detallado.columns:
            self.resumen_detallado['% Ejecutado'] = (
                self.resumen_detallado['Girado y Recaudado PAC'] / 
                self.resumen_detallado['PAC Actual'] * 100
            ).fillna(0)
        
        # Renombrar columnas para mejor visualización
        self.resumen_detallado = self.resumen_detallado.rename(columns=columnas_renombrar)
        
        # Ordenar por Período y Centro Gestor si existen
        columnas_orden = []
        if 'Período' in self.resumen_detallado.columns:
            columnas_orden.append('Período')
        if 'Centro Gestor' in self.resumen_detallado.columns:
            columnas_orden.append('Centro Gestor')
        
        if columnas_orden:
            self.resumen_detallado = self.resumen_detallado.sort_values(columnas_orden)
        
        # Resetear el índice
        self.resumen_detallado = self.resumen_detallado.reset_index(drop=True)
        
        return self.resumen_detallado
    
    def resumen_por_periodo(self):
        """Genera resumen financiero por período presupuestal"""
        self.resumen_periodo = self.df.groupby('Per.presup.').agg({
            'PAC inicial': 'sum',
            'PAC Actual': 'sum',
            'PAC Reprogramado': 'sum',
            'Girado y Recaudado PAC': 'sum',
            'Disponibilidad PAC': 'sum'
        }).reset_index()
        
        self.resumen_periodo['Ejecución %'] = (
            self.resumen_periodo['Girado y Recaudado PAC'] / 
            self.resumen_periodo['PAC Actual'] * 100
        ).fillna(0)
        
        self.resumen_periodo['Disponibilidad %'] = (
            self.resumen_periodo['Disponibilidad PAC'] / 
            self.resumen_periodo['PAC Actual'] * 100
        ).fillna(0)
        
        self.resumen_periodo['Variación PAC'] = (
            self.resumen_periodo['PAC Actual'] - 
            self.resumen_periodo['PAC inicial']
        )
        
        self.resumen_periodo['Variación %'] = (
            self.resumen_periodo['Variación PAC'] / 
            self.resumen_periodo['PAC inicial'] * 100
        ).fillna(0)
        
        return self.resumen_periodo
    
    def resumen_por_centro(self):
        """Genera resumen por centro gestor"""
        self.resumen_centro = self.df.groupby('Centro gestor').agg({
            'PAC Actual': 'sum',
            'Girado y Recaudado PAC': 'sum',
            'Disponibilidad PAC': 'sum'
        }).reset_index()
        
        self.resumen_centro['Ejecución %'] = (
            self.resumen_centro['Girado y Recaudado PAC'] / 
            self.resumen_centro['PAC Actual'] * 100
        ).fillna(0)
        
        self.resumen_centro = self.resumen_centro.sort_values('PAC Actual', ascending=False)
        return self.resumen_centro
    
    def resumen_por_fondos(self):
        """Genera resumen por tipo de fondos"""
        self.resumen_fondos = self.df.groupby('Fondos').agg({
            'PAC Actual': 'sum',
            'Girado y Recaudado PAC': 'sum',
            'Disponibilidad PAC': 'sum'
        }).reset_index()
        
        self.resumen_fondos['Ejecución %'] = (
            self.resumen_fondos['Girado y Recaudado PAC'] / 
            self.resumen_fondos['PAC Actual'] * 100
        ).fillna(0)
        
        self.resumen_fondos = self.resumen_fondos.sort_values('PAC Actual', ascending=False)
        return self.resumen_fondos
    
    def analisis_disponibilidad(self):
        """Analiza la disponibilidad presupuestal"""
        total_pac = self.df['PAC Actual'].sum()
        total_ejecutado = self.df['Girado y Recaudado PAC'].sum()
        total_disponible = self.df['Disponibilidad PAC'].sum()
        
        sobregiros = self.df[self.df['Disponibilidad PAC'] < 0]
        total_sobregiro = sobregiros['Disponibilidad PAC'].sum()
        
        self.analisis['disponibilidad'] = {
            'total_pac': total_pac,
            'total_ejecutado': total_ejecutado,
            'total_disponible': total_disponible,
            'ejecucion_pct': (total_ejecutado / total_pac * 100) if total_pac > 0 else 0,
            'disponibilidad_pct': (total_disponible / total_pac * 100) if total_pac > 0 else 0,
            'registros_sobregiro': len(sobregiros),
            'total_sobregiro': total_sobregiro
        }
        
        return self.analisis['disponibilidad']
    
    def analisis_ejecucion(self):
        """Analiza la ejecución presupuestal"""
        self.df['Ejecución %'] = (
            self.df['Girado y Recaudado PAC'] / 
            self.df['PAC Actual'] * 100
        ).fillna(0)
        
        alta = len(self.df[self.df['Ejecución %'] >= 80])
        media = len(self.df[(self.df['Ejecución %'] >= 50) & (self.df['Ejecución %'] < 80)])
        baja = len(self.df[self.df['Ejecución %'] < 50])
        
        self.analisis['ejecucion'] = {
            'alta': alta,
            'media': media,
            'baja': baja,
            'promedio_ejecucion': self.df['Ejecución %'].mean()
        }
        
        return self.analisis['ejecucion']
    
    def generar_interpretaciones(self):
        """Genera interpretaciones del análisis"""
        interpretaciones = []
        
        disp = self.analisis['disponibilidad']
        interpretaciones.append({
            'Categoría': 'RESUMEN EJECUTIVO',
            'Hallazgo': f"PAC Total: ${disp['total_pac']:,.0f}",
            'Interpretación': (
                f"El presupuesto total asignado es de ${disp['total_pac']:,.0f}. "
                f"Se ha ejecutado ${disp['total_ejecutado']:,.0f} ({disp['ejecucion_pct']:.1f}%), "
                f"quedando disponible ${disp['total_disponible']:,.0f} ({disp['disponibilidad_pct']:.1f}%)."
            ),
            'Estado': self._clasificar_ejecucion(disp['ejecucion_pct'])
        })
        
        if disp['registros_sobregiro'] > 0:
            interpretaciones.append({
                'Categoría': 'ALERTA - SOBREGIROS',
                'Hallazgo': f"{disp['registros_sobregiro']} registros con disponibilidad negativa",
                'Interpretación': (
                    f"Se detectaron {disp['registros_sobregiro']} posiciones presupuestarias con sobregiro, "
                    f"totalizando ${abs(disp['total_sobregiro']):,.0f}. Esto indica ejecución por encima "
                    f"del PAC asignado y requiere atención inmediata."
                ),
                'Estado': '⚠️ CRÍTICO'
            })
        
        if len(self.resumen_periodo) > 0:
            mejor_periodo = self.resumen_periodo.loc[
                self.resumen_periodo['Ejecución %'].idxmax()
            ]
            interpretaciones.append({
                'Categoría': 'EJECUCIÓN POR PERÍODO',
                'Hallazgo': f"Período {int(mejor_periodo['Per.presup.'])} tiene mejor ejecución",
                'Interpretación': (
                    f"El período {int(mejor_periodo['Per.presup.'])} muestra la mayor ejecución "
                    f"con {mejor_periodo['Ejecución %']:.1f}% del PAC ejecutado. "
                    f"Girado: ${mejor_periodo['Girado y Recaudado PAC']:,.0f}"
                ),
                'Estado': '✓ POSITIVO'
            })
        
        ejec = self.analisis['ejecucion']
        interpretaciones.append({
            'Categoría': 'DISTRIBUCIÓN DE EJECUCIÓN',
            'Hallazgo': f"Ejecución promedio: {ejec['promedio_ejecucion']:.1f}%",
            'Interpretación': (
                f"De {len(self.df)} registros: {ejec['alta']} tienen ejecución alta (≥80%), "
                f"{ejec['media']} media (50-80%), y {ejec['baja']} baja (<50%). "
                f"El promedio general de ejecución es {ejec['promedio_ejecucion']:.1f}%."
            ),
            'Estado': self._clasificar_ejecucion(ejec['promedio_ejecucion'])
        })
        
        if len(self.resumen_centro) >= 3:
            top3 = self.resumen_centro.head(3)
            interpretaciones.append({
                'Categoría': 'PRINCIPALES CENTROS GESTORES',
                'Hallazgo': f"Top 3 concentran ${top3['PAC Actual'].sum():,.0f}",
                'Interpretación': (
                    f"Los 3 principales centros gestores son: "
                    f"{', '.join(top3['Centro gestor'].tolist())}. "
                    f"Juntos representan el {(top3['PAC Actual'].sum() / disp['total_pac'] * 100):.1f}% "
                    f"del PAC total."
                ),
                'Estado': '📊 INFO'
            })
        
        return pd.DataFrame(interpretaciones)
    
    def _clasificar_ejecucion(self, porcentaje):
        """Clasifica el nivel de ejecución"""
        if porcentaje >= 80:
            return '✓ EXCELENTE'
        elif porcentaje >= 60:
            return '✓ BUENO'
        elif porcentaje >= 40:
            return '⚠️ REGULAR'
        else:
            return '❌ BAJO'
    
    def exportar_excel(self):
        """Exporta todos los análisis a Excel"""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Estilos
        header_fill = PatternFill(start_color='DA121A', end_color='DA121A', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        title_font = Font(bold=True, size=14, color='DA121A')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Hoja 1: Resumen Ejecutivo
        ws1 = wb.create_sheet('Resumen Ejecutivo')
        ws1['A1'] = 'ANÁLISIS COMPLETO DEL REPORTE PAC'
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:D1')
        
        ws1['A3'] = 'INDICADORES GENERALES'
        ws1['A3'].font = Font(bold=True, size=12)
        
        disp = self.analisis['disponibilidad']
        metricas = [
            ['Métrica', 'Valor', 'Unidad', 'Estado'],
            ['PAC Total', disp['total_pac'], '$', ''],
            ['Girado y Recaudado', disp['total_ejecutado'], '$', ''],
            ['Disponibilidad', disp['total_disponible'], '$', ''],
            ['% Ejecución', disp['ejecucion_pct'], '%', self._clasificar_ejecucion(disp['ejecucion_pct'])],
            ['% Disponibilidad', disp['disponibilidad_pct'], '%', ''],
            ['Registros con Sobregiro', disp['registros_sobregiro'], 'unidades', 
             '⚠️ CRÍTICO' if disp['registros_sobregiro'] > 0 else '✓ OK']
        ]
        
        for row_idx, row_data in enumerate(metricas, start=4):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border
                if col_idx == 2 and row_idx > 4:
                    if row_data[2] == '$':
                        cell.number_format = '$#,##0'
                    elif row_data[2] == '%':
                        cell.number_format = '0.0'
        
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 15
        
        # Hoja 2: Análisis Detallado
        ws2 = wb.create_sheet('Análisis Detallado')
        ws2['A1'] = 'ANÁLISIS DETALLADO COMPLETO'
        ws2['A1'].font = title_font
        
        if self.resumen_detallado is not None and len(self.resumen_detallado) > 0:
            num_cols = len(self.resumen_detallado.columns)
            ws2.merge_cells(f'A1:{chr(64 + num_cols)}1')
            
            for r_idx, row in enumerate(dataframe_to_rows(self.resumen_detallado, index=False, header=True), start=3):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 3:
                        cell.fill = header_fill
                        cell.font = header_font
                    cell.border = border
                    
                    # Formatear números
                    if r_idx > 3:
                        col_name = self.resumen_detallado.columns[c_idx - 1]
                        if 'PAC' in col_name or 'Girado' in col_name or 'Disponibilidad' in col_name:
                            cell.number_format = '$#,##0'
                        elif '%' in col_name or 'Ejecutado' in col_name:
                            cell.number_format = '0.00'
            
            # Ajustar anchos
            for idx, col in enumerate(self.resumen_detallado.columns, start=1):
                col_letter = chr(64 + idx)
                if 'Pos.Presupuestaria' in col or 'Programa' in col:
                    ws2.column_dimensions[col_letter].width = 25
                elif 'Centro' in col:
                    ws2.column_dimensions[col_letter].width = 20
                else:
                    ws2.column_dimensions[col_letter].width = 18
        
        # Hoja 3: Por Período
        ws3 = wb.create_sheet('Por Período')
        ws3['A1'] = 'ANÁLISIS POR PERÍODO PRESUPUESTAL'
        ws3['A1'].font = title_font
        ws3.merge_cells('A1:I1')
        
        for r_idx, row in enumerate(dataframe_to_rows(self.resumen_periodo, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border
                
                if r_idx > 3 and c_idx >= 2:
                    if c_idx in [7, 8, 10]:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '$#,##0'
        
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws3.column_dimensions[col_letter].width = 18
        
        # Hoja 4: Por Centro Gestor
        ws4 = wb.create_sheet('Por Centro Gestor')
        ws4['A1'] = 'ANÁLISIS POR CENTRO GESTOR'
        ws4['A1'].font = title_font
        ws4.merge_cells('A1:E1')
        
        for r_idx, row in enumerate(dataframe_to_rows(self.resumen_centro, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border
                
                if r_idx > 3 and c_idx >= 2:
                    if c_idx == 5:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '$#,##0'
        
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws4.column_dimensions[col_letter].width = 20
        
        # Hoja 5: Por Fondos
        ws5 = wb.create_sheet('Por Fondos')
        ws5['A1'] = 'ANÁLISIS POR TIPO DE FONDOS'
        ws5['A1'].font = title_font
        ws5.merge_cells('A1:E1')
        
        for r_idx, row in enumerate(dataframe_to_rows(self.resumen_fondos, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws5.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border
                
                if r_idx > 3 and c_idx >= 2:
                    if c_idx == 5:
                        cell.number_format = '0.0'
                    else:
                        cell.number_format = '$#,##0'
        
        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            ws5.column_dimensions[col_letter].width = 20
        
        # Hoja 6: Interpretaciones
        interpretaciones_df = self.generar_interpretaciones()
        ws6 = wb.create_sheet('Interpretaciones')
        ws6['A1'] = 'INTERPRETACIÓN Y HALLAZGOS DEL ANÁLISIS'
        ws6['A1'].font = title_font
        ws6.merge_cells('A1:D1')
        
        for r_idx, row in enumerate(dataframe_to_rows(interpretaciones_df, index=False, header=True), start=3):
            for c_idx, value in enumerate(row, start=1):
                cell = ws6.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 3:
                    cell.fill = header_fill
                    cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        ws6.column_dimensions['A'].width = 20
        ws6.column_dimensions['B'].width = 35
        ws6.column_dimensions['C'].width = 60
        ws6.column_dimensions['D'].width = 15
        
        for row in ws6.iter_rows(min_row=4, max_row=ws6.max_row):
            ws6.row_dimensions[row[0].row].height = 45
        
        return wb

# ============================================================
# FUNCIONES DE AUTENTICACIÓN
# ============================================================

def validar_login(user: str, password: str) -> Tuple[bool, Optional[Dict]]:
    """Valida las credenciales del usuario"""
    user = user.strip().lower()
    if user in USUARIOS:
        if USUARIOS[user]["password"] == password:
            return True, USUARIOS[user]
    return False, None

def tiene_permiso(permiso: str) -> bool:
    """Verifica si el usuario actual tiene un permiso específico"""
    if not st.session_state.logged_in:
        return False
    return permiso in st.session_state.get("permisos", [])

# ============================================================
# CONFIGURACIÓN DE STREAMLIT
# ============================================================

st.set_page_config(
    page_title="Alcaldía Local de Usme - Análisis PAC",
    page_icon="🏛️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================
# ESTILOS CSS
# ============================================================

st.markdown(f"""
    <style>
    .main {{
        background-color: {WHITE};
    }}
    
    [data-testid="stSidebar"] {{
        background: linear-gradient(180deg, {BOGOTA_YELLOW} 0%, {BOGOTA_RED} 100%);
    }}
    
    [data-testid="stSidebar"] * {{
        color: {BLACK} !important;
    }}
    
    h1 {{
        color: {BOGOTA_RED};
        font-weight: 700;
        border-bottom: 3px solid {BOGOTA_YELLOW};
        padding-bottom: 10px;
    }}
    
    h2, h3 {{
        color: {DARK};
    }}
    
    .stButton > button {{
        background-color: {BOGOTA_RED};
        color: {WHITE};
        border: none;
        border-radius: 5px;
        padding: 10px 20px;
        font-weight: 600;
        transition: all 0.3s;
    }}
    
    .stButton > button:hover {{
        background-color: {DARK};
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(0,0,0,0.2);
    }}
    
    .stTabs [data-baseweb="tab-list"] {{
        gap: 10px;
    }}
    
    .stTabs [data-baseweb="tab"] {{
        background-color: {BOGOTA_YELLOW};
        color: {BLACK};
        border-radius: 5px 5px 0 0;
        padding: 10px 20px;
        font-weight: 600;
    }}
    
    .stTabs [aria-selected="true"] {{
        background-color: {BOGOTA_RED};
        color: {WHITE};
    }}
    
    .logo-box {{
        background-color: {WHITE};
        border-radius: 10px;
        padding: 20px;
        text-align: center;
        margin-bottom: 20px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }}
    
    .logo-text {{
        font-size: 32px;
        font-weight: 900;
        color: {BOGOTA_RED};
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }}
    
    [data-testid="stMetricValue"] {{
        color: {BOGOTA_RED};
        font-size: 24px;
        font-weight: 700;
    }}
    
    hr {{
        border: none;
        height: 2px;
        background: linear-gradient(90deg, {BOGOTA_YELLOW} 0%, {BOGOTA_RED} 100%);
        margin: 20px 0;
    }}
    
    .info-card {{
        background-color: {WHITE};
        border-left: 4px solid {BOGOTA_RED};
        padding: 15px;
        margin: 10px 0;
        border-radius: 5px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }}
    </style>
""", unsafe_allow_html=True)

# ============================================================
# INICIALIZACIÓN DE SESSION STATE
# ============================================================

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "user" not in st.session_state:
    st.session_state.user = ""
if "rol" not in st.session_state:
    st.session_state.rol = ""
if "permisos" not in st.session_state:
    st.session_state.permisos = []
if "failed_attempts" not in st.session_state:
    st.session_state.failed_attempts = 0
if "lock_until" not in st.session_state:
    st.session_state.lock_until = None
if "historial_accesos" not in st.session_state:
    st.session_state.historial_accesos = []

# ============================================================
# SIDEBAR - LOGIN Y CONTROL
# ============================================================

with st.sidebar:
    st.markdown(
        f'<div class="logo-box"><div class="logo-text">🏛️ USME</div></div>',
        unsafe_allow_html=True
    )
    
    st.markdown('### 🔐 Control de Acceso')
    
    if st.session_state.lock_until is not None:
        ahora = datetime.now()
        if ahora < st.session_state.lock_until:
            restante = int((st.session_state.lock_until - ahora).total_seconds())
            st.error(f"🔒 **Acceso bloqueado**\n\nIntenta de nuevo en {restante} segundos")
            st.stop()
        else:
            st.session_state.lock_until = None
            st.session_state.failed_attempts = 0
    
    if not st.session_state.logged_in:
        with st.form("login_form"):
            st.markdown("#### Iniciar Sesión")
            usuario = st.text_input("👤 Usuario", placeholder="Ingresa tu usuario")
            clave = st.text_input("🔑 Contraseña", type="password", placeholder="Ingresa tu contraseña")
            
            if MOSTRAR_INTENTOS and st.session_state.failed_attempts > 0:
                st.warning(f"⚠️ Intento {st.session_state.failed_attempts} de {MAX_INTENTOS}")
            
            col1, col2 = st.columns(2)
            login_btn = col1.form_submit_button("🚀 Ingresar", use_container_width=True)
            help_btn = col2.form_submit_button("❓ Ayuda", use_container_width=True)
            
            if help_btn:
                st.info("""
                **Usuarios disponibles:**
                - admin / admin123
                - contador / contador2025
                - auxiliar / aux123
                - auditor / audit2025
                - jefe / jefe2025
                """)
            
            if login_btn:
                exito, datos_usuario = validar_login(usuario, clave)
                
                if exito:
                    st.session_state.logged_in = True
                    st.session_state.user = usuario.strip().lower()
                    st.session_state.rol = datos_usuario["rol"]
                    st.session_state.permisos = datos_usuario["permisos"]
                    st.session_state.failed_attempts = 0
                    
                    st.session_state.historial_accesos.append({
                        "usuario": st.session_state.user,
                        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "resultado": "Exitoso"
                    })
                    
                    st.success(f"✅ Bienvenido {datos_usuario['rol']}")
                    st.rerun()
                else:
                    st.session_state.failed_attempts += 1
                    restantes = MAX_INTENTOS - st.session_state.failed_attempts
                    
                    st.session_state.historial_accesos.append({
                        "usuario": usuario,
                        "fecha": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "resultado": "Fallido"
                    })
                    
                    if restantes > 0:
                        st.error(f"❌ Credenciales incorrectas\n\n{restantes} intento(s) restante(s)")
                    else:
                        st.error("❌ Has agotado los intentos permitidos")
                        st.session_state.lock_until = datetime.now() + timedelta(minutes=BLOQUEO_MINUTOS)
                        st.warning(f"🔒 Bloqueado por {BLOQUEO_MINUTOS} minuto(s)")
                        st.session_state.failed_attempts = 0
                    
                    st.rerun()
    
    else:
        st.success(f"✅ **Sesión Activa**")
        st.markdown(f"""
        <div style='background-color: white; padding: 15px; border-radius: 10px; margin: 10px 0;'>
            <p style='margin: 0; color: black;'><strong>👤 Usuario:</strong> {st.session_state.user}</p>
            <p style='margin: 5px 0 0 0; color: black;'><strong>👔 Rol:</strong> {st.session_state.rol}</p>
        </div>
        """, unsafe_allow_html=True)
        
        if st.button("🚪 Cerrar Sesión", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.user = ""
            st.session_state.rol = ""
            st.session_state.permisos = []
            st.session_state.failed_attempts = 0
            st.session_state.lock_until = None
            st.rerun()
        
        st.markdown("---")
        
        st.markdown("### 📊 Información del Sistema")
        
        st.markdown("#### 🔔 Alertas")
        col1, col2 = st.columns(2)
        col1.checkbox("🔒 Seguridad", value=True, disabled=True)
        col2.checkbox("🌐 Internet", value=True, disabled=True)
        
        st.markdown("#### 📈 Métricas")
        m1, m2 = st.columns(2)
        m1.metric("Usuarios", len(USUARIOS))
        m2.metric("Estado", "🟢 Activo")
        
        if tiene_permiso("analisis_pac"):
            st.markdown("#### 📋 Historial")
            with st.expander("Ver últimos accesos"):
                if st.session_state.historial_accesos:
                    for acceso in st.session_state.historial_accesos[-5:]:
                        resultado_icon = "✅" if acceso["resultado"] == "Exitoso" else "❌"
                        st.caption(f"{resultado_icon} {acceso['usuario']} - {acceso['fecha']}")
                else:
                    st.caption("No hay historial disponible")

# ============================================================
# CONTENIDO PRINCIPAL
# ============================================================

st.title("🏛️ Alcaldía Local de Usme")
st.caption("Sistema de Análisis PAC • Plan Anual de Caja")

if st.session_state.logged_in:
    col1, col2, col3, col4 = st.columns(4)
    col1.info(f"👤 **Usuario:** {st.session_state.user}")
    col2.info(f"👔 **Rol:** {st.session_state.rol}")
    col3.info(f"📅 **Fecha:** {datetime.now().strftime('%d/%m/%Y')}")
    col4.info(f"🕐 **Hora:** {datetime.now().strftime('%H:%M:%S')}")

st.markdown("---")

# ============================================================
# CONTENIDO PRINCIPAL - ANÁLISIS PAC
# ============================================================

if not st.session_state.logged_in:
    st.warning("🔒 **Acceso Restringido**")
    st.info("Por favor, inicia sesión desde el panel izquierdo para acceder al sistema de análisis PAC.")
    
    st.markdown("""
    <div class="info-card">
        <h4>📊 Sistema de Análisis PAC</h4>
        <p>Este sistema permite analizar reportes del Plan Anual de Caja con:</p>
        <ul>
            <li>📈 Análisis detallado por múltiples dimensiones</li>
            <li>💰 Métricas de ejecución presupuestal</li>
            <li>⚠️ Detección de sobregiros</li>
            <li>📊 Reportes en Excel profesionales</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)

elif not tiene_permiso("analisis_pac"):
    st.error("❌ **Acceso Denegado**")
    st.warning(f"Tu rol ({st.session_state.rol}) no tiene permisos para análisis PAC.")

else:
    st.subheader("📊 Análisis Completo de Reportes PAC")
    st.success("✅ **Acceso Autorizado** - Puedes analizar reportes PAC")
    
    st.markdown("""
    <div class="info-card">
        <h4>📋 Análisis Disponibles</h4>
        <ul>
            <li>📊 <strong>Análisis Detallado</strong> - Por programa, centro, período, posición y fondos</li>
            <li>📅 <strong>Por Período</strong> - Ejecución presupuestal por período</li>
            <li>🏢 <strong>Por Centro Gestor</strong> - Distribución por centros</li>
            <li>💰 <strong>Por Fondos</strong> - Análisis por tipo de fondos</li>
            <li>🔍 <strong>Interpretaciones</strong> - Hallazgos automáticos</li>
        </ul>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown("---")
    
    # Uploader
    archivo_pac = st.file_uploader(
        "📤 Subir Archivo Reporte PAC",
        type=["xlsx", "xls"],
        help="Sube el archivo Excel del reporte PAC"
    )
    
    if archivo_pac is not None:
        try:
            # Leer archivo
            with st.spinner("📂 Cargando datos..."):
                df = pd.read_excel(archivo_pac, sheet_name='Data')
            
            st.success(f"✅ Archivo cargado: {len(df)} registros encontrados")
            
            # Mostrar preview
            with st.expander("👁️ Vista previa de datos (primeras 10 filas)"):
                st.dataframe(df.head(10), use_container_width=True)
            
            # Botón de análisis
            if st.button("🚀 Iniciar Análisis Completo", type="primary", use_container_width=True):
                with st.spinner("⚡ Procesando análisis... Por favor espera."):
                    try:
                        # Crear analizador
                        analizador = AnalizadorPAC(df)
                        analizador.limpiar_datos()
                        
                        # Progress bar
                        progress_bar = st.progress(0)
                        status_text = st.empty()
                        
                        # Ejecutar análisis
                        status_text.text("📊 Generando análisis detallado...")
                        analizador.resumen_detallado_completo()
                        progress_bar.progress(15)
                        
                        status_text.text("📅 Analizando por período...")
                        analizador.resumen_por_periodo()
                        progress_bar.progress(35)
                        
                        status_text.text("🏢 Analizando por centro gestor...")
                        analizador.resumen_por_centro()
                        progress_bar.progress(55)
                        
                        status_text.text("💰 Analizando por fondos...")
                        analizador.resumen_por_fondos()
                        progress_bar.progress(70)
                        
                        status_text.text("📈 Calculando disponibilidad...")
                        analizador.analisis_disponibilidad()
                        progress_bar.progress(85)
                        
                        status_text.text("⚡ Analizando ejecución...")
                        analizador.analisis_ejecucion()
                        progress_bar.progress(95)
                        
                        status_text.text("🔍 Generando interpretaciones...")
                        interpretaciones = analizador.generar_interpretaciones()
                        progress_bar.progress(100)
                        
                        status_text.text("✅ ¡Análisis completado!")
                        
                        st.success("🎉 **¡Análisis Completado Exitosamente!**")
                        
                        # Mostrar métricas principales
                        st.markdown("### 📊 Métricas Principales")
                        
                        disp = analizador.analisis['disponibilidad']
                        
                        col1, col2, col3, col4 = st.columns(4)
                        col1.metric(
                            "💰 PAC Total",
                            f"${disp['total_pac']:,.0f}"
                        )
                        col2.metric(
                            "✅ Ejecutado",
                            f"${disp['total_ejecutado']:,.0f}",
                            f"{disp['ejecucion_pct']:.1f}%"
                        )
                        col3.metric(
                            "📊 Disponible",
                            f"${disp['total_disponible']:,.0f}",
                            f"{disp['disponibilidad_pct']:.1f}%"
                        )
                        col4.metric(
                            "⚠️ Sobregiros",
                            disp['registros_sobregiro'],
                            "CRÍTICO" if disp['registros_sobregiro'] > 0 else "OK"
                        )
                        
                        st.markdown("---")
                        
                        # Tabs de resultados
                        result_tab1, result_tab2, result_tab3, result_tab4, result_tab5 = st.tabs([
                            "📊 Análisis Detallado", "📅 Por Período", "🏢 Por Centro", "💰 Por Fondos", "🔍 Interpretaciones"
                        ])
                        
                        with result_tab1:
                            st.markdown("#### 📊 Análisis Detallado Completo")
                            if analizador.resumen_detallado is not None and len(analizador.resumen_detallado) > 0:
                                # Formatear columnas para visualización
                                formato_dict = {}
                                for col in analizador.resumen_detallado.columns:
                                    if 'PAC' in col or 'Girado' in col or 'Disponibilidad' in col:
                                        formato_dict[col] = '${:,.0f}'
                                    elif 'Ejecutado' in col or '%' in col:
                                        formato_dict[col] = '{:.2f}%'
                                
                                st.dataframe(
                                    analizador.resumen_detallado.style.format(formato_dict),
                                    use_container_width=True,
                                    height=600
                                )
                                
                                st.info(f"📊 Total de registros: {len(analizador.resumen_detallado)}")
                            else:
                                st.warning("⚠️ No se pudo generar el análisis detallado")
                        
                        with result_tab2:
                            st.markdown("#### 📅 Resumen por Período Presupuestal")
                            st.dataframe(
                                analizador.resumen_periodo.style.format({
                                    'PAC inicial': '${:,.0f}',
                                    'PAC Actual': '${:,.0f}',
                                    'PAC Reprogramado': '${:,.0f}',
                                    'Girado y Recaudado PAC': '${:,.0f}',
                                    'Disponibilidad PAC': '${:,.0f}',
                                    'Ejecución %': '{:.1f}%',
                                    'Disponibilidad %': '{:.1f}%',
                                    'Variación PAC': '${:,.0f}',
                                    'Variación %': '{:.1f}%'
                                }),
                                use_container_width=True
                            )
                        
                        with result_tab3:
                            st.markdown("#### 🏢 Resumen por Centro Gestor")
                            st.dataframe(
                                analizador.resumen_centro.head(20).style.format({
                                    'PAC Actual': '${:,.0f}',
                                    'Girado y Recaudado PAC': '${:,.0f}',
                                    'Disponibilidad PAC': '${:,.0f}',
                                    'Ejecución %': '{:.1f}%'
                                }),
                                use_container_width=True
                            )
                        
                        with result_tab4:
                            st.markdown("#### 💰 Resumen por Tipo de Fondos")
                            st.dataframe(
                                analizador.resumen_fondos.style.format({
                                    'PAC Actual': '${:,.0f}',
                                    'Girado y Recaudado PAC': '${:,.0f}',
                                    'Disponibilidad PAC': '${:,.0f}',
                                    'Ejecución %': '{:.1f}%'
                                }),
                                use_container_width=True
                            )
                        
                        with result_tab5:
                            st.markdown("#### 🔍 Interpretaciones y Hallazgos")
                            for idx, row in interpretaciones.iterrows():
                                with st.expander(f"{row['Estado']} | {row['Categoría']}"):
                                    st.markdown(f"**Hallazgo:** {row['Hallazgo']}")
                                    st.markdown(f"**Interpretación:** {row['Interpretación']}")
                        
                        st.markdown("---")
                        
                        # Generar Excel
                        st.markdown("### 📥 Descargar Reporte Completo")
                        
                        with st.spinner("📊 Generando archivo Excel..."):
                            wb = analizador.exportar_excel()
                            
                            # Guardar en bytes
                            from io import BytesIO
                            buffer = BytesIO()
                            wb.save(buffer)
                            buffer.seek(0)
                            
                            fecha_actual = datetime.now().strftime("%Y%m%d_%H%M%S")
                            nombre_archivo = f"ANALISIS_PAC_{fecha_actual}.xlsx"
                            
                            st.download_button(
                                label="⬇️ Descargar Análisis Completo (Excel)",
                                data=buffer,
                                file_name=nombre_archivo,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                type="primary"
                            )
                        
                        st.success("✅ Archivo Excel generado exitosamente")
                        st.info("📋 El archivo contiene 6 hojas: Resumen Ejecutivo, Análisis Detallado, Por Período, Por Centro Gestor, Por Fondos e Interpretaciones")
                        st.balloons()
                    
                    except Exception as e:
                        st.error("❌ **Error durante el análisis**")
                        st.exception(e)
        
        except Exception as e:
            st.error("❌ **Error al cargar el archivo**")
            st.exception(e)

# ============================================================
# FOOTER
# ============================================================

st.markdown("---")
st.markdown("""
<div style='text-align: center; color: gray; padding: 20px;'>
    <p>🏛️ <strong>Alcaldía Local de Usme</strong> • Sistema de Análisis PAC</p>
    <p>📍 Bogotá D.C., Colombia • 📞 Contacto: (320) 830-38-47</p>
    <p style='font-size: 12px;'>© 2026 Richard Bello Roncancio. Todos los derechos reservados.</p>
</div>
""", unsafe_allow_html=True)
